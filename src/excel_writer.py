from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import logging
import pandas as pd

logger = logging.getLogger(__name__)

def generar_reporte(tablas, indices, config):
    """Genera el reporte Excel escribiendo todas las tablas y datos en la plantilla."""
    try:
        template = config.TEMPLATES / f"Cuadros de RMC-{config.MES_ACTUAL}-{config.ANIOS[0]}-Joel-act.xlsx"
        output = config.OUTPUT_REPORTES / f"RMC_{config.MES_ACTUAL}_{config.ANIOS[0]}.xlsx"
        
        if not template.exists():
            raise FileNotFoundError(f"Plantilla no encontrada: {template}")
        
        config.OUTPUT_REPORTES.mkdir(parents=True, exist_ok=True)
        
        logger.info(f"Cargando plantilla: {template}")
        libro = load_workbook(template)
        libro.calculation.calcMode = "auto"
        libro.calculation.fullCalcOnLoad = True
        libro.calculation.forceFullCalc = True

        # LLENAR HOJA COMERCIO_TEXTIL
        logger.info("Escribiendo datos en hoja Comercio_Textil...")
        tabla_final = tablas['tabla_final']
        detalle_textil = tablas.get('detalle_textil', pd.DataFrame())
        detalle_textil_import = tablas.get('detalle_textil_import', pd.DataFrame())
        tabla_destinos = tablas['tabla_destinos']
        num_destinos = tablas['num_destinos']
        
        _escribir_comercio_textil(
            libro['Comercio_Textil'],
            tabla_final,
            detalle_textil,
            detalle_textil_import,
            tabla_destinos,
            num_destinos,
        )
        _escribir_indices_textil(libro['Indices_X_Textil'], indices.get('indice_textil', {}))
        _escribir_indices_textil_import(libro['Indices_M_Textil'], indices.get('indice_textil_import', {}))
        
        logger.info(f"Guardando reporte en: {output}")
        libro.save(output)
        logger.info("Reporte completado exitosamente")
        
    except FileNotFoundError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Error generando reporte: {str(e)}")
        raise


def _escribir_comercio_agro(hoja, tabla_final, tabla_destinos, num_destinos):
    """Escribe datos en la hoja Comercio_Agro."""
    try:
        # Flujos y grupos de productos
        for t in range(0, 3):
            hoja.cell(10, 8+t).value = tabla_final.iloc[0, t]
        for t in range(0, 2):
            hoja.cell(10, 13+t).value = tabla_final.iloc[0, t+4]
        
        # Grupos
        for x in range(0, 2):
            hoja.cell(11+(15*x), 6).value = _index_label(tabla_final.index[x+1])
            for t in range(0, 3):
                hoja.cell(11+(15*x), 8+t).value = tabla_final.iloc[x+1, t]
            for t in range(0, 2):
                hoja.cell(11+(15*x), 13+t).value = tabla_final.iloc[x+1, t+4]
        
        # Principales destinos
        for x in range(0, 5):
            hoja.cell(55+x, 6).value = _index_label(tabla_destinos.index[x+1])
            for t in range(0, 3):
                hoja.cell(55+x, 8+t).value = tabla_destinos.iloc[x+1, t]
            hoja.cell(55+x, 13).value = tabla_destinos.iloc[x+1, 4]
            hoja.cell(55+x, 14).value = tabla_destinos.iloc[x+1, 5]
        
        # Numero de destinos
        for t in range(0, 3):
            hoja.cell(62, 8+t).value = num_destinos.iloc[0, t]
        
        logger.debug("Hoja Comercio_Agro completada")
    except Exception as e:
        logger.error(f"Error escribiendo Comercio_Agro: {str(e)}")
        raise


def _index_label(value):
    if isinstance(value, tuple):
        for item in reversed(value):
            if item not in (None, ''):
                return str(item)
        return ''
    return str(value)


def _escribir_comercio_textil(hoja, tabla_final, detalle_textil, detalle_textil_import, tabla_destinos, num_destinos):
    """Escribe datos en la hoja Comercio_Textil."""
    try:
        _escribir_comercio_textil_exportaciones(hoja, tabla_final, detalle_textil, tabla_destinos, num_destinos)
        _escribir_comercio_textil_importaciones(hoja, detalle_textil_import)
        
        logger.debug("Hoja Comercio_Textil completada")
    except Exception as e:
        logger.error(f"Error escribiendo Comercio_Textil: {str(e)}")
        raise


def _escribir_comercio_textil_exportaciones(hoja, tabla_final, detalle_textil, tabla_destinos, num_destinos):
    """Escribe la seccion de exportaciones en Comercio_Textil."""
    row_map = {0: 10, 1: 12, 2: 18}
    for idx, row in row_map.items():
        for t in range(0, 3):
            hoja.cell(row, 8+t).value = tabla_final.iloc[idx, t]
        for t in range(0, 2):
            hoja.cell(row, 13+t).value = tabla_final.iloc[idx, t+4]

    detalle_map = {
        'prendas_vestir': 13,
        'prendas_algodon': 14,
        'mantas_pelo_fino': 16,
        'mantas_algodon': 17,
        'fibras_textiles': 19,
        'tejidos': 20,
        'tejidos_algodon': 21,
        'hilos_hilados': 22,
    }
    for etiqueta, row in detalle_map.items():
        _escribir_resumen_fila(hoja, row, detalle_textil, etiqueta)
        
    for x in range(0, 3):
        hoja.cell(25+x, 6).value = _index_label(tabla_destinos.index[x+1])
        for t in range(0, 3):
            hoja.cell(25+x, 8+t).value = tabla_destinos.iloc[x+1, t]
        hoja.cell(25+x, 13).value = tabla_destinos.iloc[x+1, 4]
        hoja.cell(25+x, 14).value = tabla_destinos.iloc[x+1, 5]

    for t in range(0, 3):
        hoja.cell(28, 8+t).value = num_destinos.iloc[0, t]


def _escribir_comercio_textil_importaciones(hoja, detalle_textil_import):
    """Escribe la seccion de importaciones en Comercio_Textil."""
    _escribir_resumen_fila(hoja, 29, detalle_textil_import, 'sector_total')

    bloques = [
        ('textiles', _valor_ult12(detalle_textil_import, 'textiles')),
        ('confecciones', _valor_ult12(detalle_textil_import, 'confecciones')),
    ]
    bloques.sort(key=lambda item: item[1], reverse=True)

    block_rows = [31, 40]
    for row_start, (bloque, _) in zip(block_rows, bloques):
        if bloque == 'textiles':
            _escribir_bloque_import_textiles(hoja, detalle_textil_import, row_start)
        else:
            _escribir_bloque_import_confecciones(hoja, detalle_textil_import, row_start)


def _escribir_bloque_import_textiles(hoja, tabla_resumen, row_start):
    _limpiar_area_importaciones(hoja, row_start, row_start + 8)

    resumen_row = row_start
    _escribir_resumen_fila(hoja, resumen_row, tabla_resumen, 'textiles')
    hoja.cell(resumen_row, 6).value = 'Textiles [2]'
    _escribir_indice_referencia(hoja, resumen_row, 'textiles', 'Textiles')

    _escribir_resumen_fila(hoja, row_start + 1, tabla_resumen, 'tejidos')
    hoja.cell(row_start + 1, 6).value = 'Tejidos'
    _escribir_indice_referencia(hoja, row_start + 1, 'tejidos', 'Tejidos')

    tejidos_subs = [
        ('tejidos_poliester', '   -  De poliester'),
        ('tejidos_algodon', '   -  De algodón'),
    ]
    tejidos_subs.sort(key=lambda item: _valor_ult12(tabla_resumen, item[0]), reverse=True)
    for offset, (etiqueta, label) in enumerate(tejidos_subs, start=2):
        _escribir_resumen_fila(hoja, row_start + offset, tabla_resumen, etiqueta)
        hoja.cell(row_start + offset, 6).value = label

    _escribir_resumen_fila(hoja, row_start + 4, tabla_resumen, 'hilos_hilados')
    hoja.cell(row_start + 4, 6).value = 'Hilos e hilados'
    _escribir_indice_referencia(hoja, row_start + 4, 'hilos_hilados', 'Hilos e hilados')

    _escribir_resumen_fila(hoja, row_start + 5, tabla_resumen, 'hilos_algodon')
    hoja.cell(row_start + 5, 6).value = '   -  De algodón'

    _escribir_resumen_fila(hoja, row_start + 6, tabla_resumen, 'fibras_textiles')
    hoja.cell(row_start + 6, 6).value = 'Fibras textiles'
    _escribir_indice_referencia(hoja, row_start + 6, 'fibras_textiles', 'Fibras textiles')

    if row_start + 7 <= 47:
        _limpiar_fila_importacion(hoja, row_start + 7)
    if row_start + 8 <= 48:
        _limpiar_fila_importacion(hoja, row_start + 8)


def _escribir_bloque_import_confecciones(hoja, tabla_resumen, row_start):
    end_row = min(row_start + 7, 47)
    _limpiar_area_importaciones(hoja, row_start, end_row)

    _escribir_resumen_fila(hoja, row_start, tabla_resumen, 'confecciones')
    hoja.cell(row_start, 6).value = 'Confecciones [2]'
    _escribir_indice_referencia(hoja, row_start, 'confecciones', 'Confecciones')

    padres = [
        ('prendas_vestir', 'Prendas de vestir', [('prendas_algodon', '   -  De algodón'), ('prendas_sinteticas', '   -  Sintéticas')]),
        ('otras_confecciones', 'Otras confecciones', [('mantas_fibra_sintetica', '   -  Mantas de fibra sintética'), ('ropa_de_cama', '   -  Ropa de cama')]),
    ]
    padres.sort(key=lambda item: _valor_ult12(tabla_resumen, item[0]), reverse=True)

    row_pairs = [(row_start + 1, [row_start + 2, row_start + 3]), (row_start + 4, [row_start + 5, row_start + 6])]
    for (etiqueta, label, hijos), (parent_row, child_rows) in zip(padres, row_pairs):
        _escribir_resumen_fila(hoja, parent_row, tabla_resumen, etiqueta)
        hoja.cell(parent_row, 6).value = label
        _escribir_indice_referencia(hoja, parent_row, etiqueta, label.strip())

        hijos_ordenados = sorted(hijos, key=lambda item: _valor_ult12(tabla_resumen, item[0]), reverse=True)
        for child_row, (child_etiqueta, child_label) in zip(child_rows, hijos_ordenados):
            _escribir_resumen_fila(hoja, child_row, tabla_resumen, child_etiqueta)
            hoja.cell(child_row, 6).value = child_label

    if row_start + 7 <= 47:
        _limpiar_fila_importacion(hoja, row_start + 7)


def _valor_ult12(tabla_resumen, etiqueta):
    if tabla_resumen is None or tabla_resumen.empty or etiqueta not in tabla_resumen.index:
        return float('-inf')
    return tabla_resumen.loc[etiqueta].iloc[3]


def _escribir_indice_referencia(hoja, row, etiqueta, label):
    referencias = {
        'sector_total': ('H', 'M'),
        'textiles': ('AA', 'AF'),
        'tejidos': ('AT', 'AY'),
        'hilos_hilados': ('BM', 'BR'),
        'fibras_textiles': ('CF', 'CK'),
        'confecciones': ('CY', 'DD'),
        'prendas_vestir': ('DR', 'DW'),
        'otras_confecciones': ('EK', 'EP'),
    }

    hoja.cell(row, 22).value = label
    cols = referencias.get(etiqueta)
    if not cols:
        for col in range(23, 29):
            _set_safe_value(hoja, row, col, None)
        return

    inicio, fin = cols
    letras = _column_range(inicio, fin)
    for idx, col_letter in enumerate(letras, start=23):
        hoja.cell(row, idx).value = f"=Indices_M_Textil!{col_letter}6"


def _column_range(start, end):
    from openpyxl.utils import column_index_from_string, get_column_letter

    start_idx = column_index_from_string(start)
    end_idx = column_index_from_string(end)
    return [get_column_letter(idx) for idx in range(start_idx, end_idx + 1)]


def _limpiar_area_importaciones(hoja, row_start, row_end):
    for row in range(row_start, row_end + 1):
        _limpiar_fila_importacion(hoja, row)


def _limpiar_fila_importacion(hoja, row):
    _set_safe_value(hoja, row, 6, None)
    _set_safe_value(hoja, row, 22, None)
    for col in range(8, 11):
        _set_safe_value(hoja, row, col, None)
    for col in range(13, 15):
        _set_safe_value(hoja, row, col, None)
    for col in range(23, 29):
        _set_safe_value(hoja, row, col, None)


def _set_safe_value(hoja, row, col, value):
    cell = hoja.cell(row, col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def _escribir_resumen_fila(hoja, row, tabla_resumen, etiqueta):
    for col in range(8, 11):
        hoja.cell(row, col).value = None
    for col in range(13, 15):
        hoja.cell(row, col).value = None

    if tabla_resumen is None or tabla_resumen.empty or etiqueta not in tabla_resumen.index:
        return

    for t in range(0, 3):
        hoja.cell(row, 8+t).value = tabla_resumen.loc[etiqueta].iloc[t]
    for t in range(0, 2):
        hoja.cell(row, 13+t).value = tabla_resumen.loc[etiqueta].iloc[t+4]


def _escribir_comercio_pesca(hoja, tabla_final, tabla_destinos, num_destinos):
    """Escribe datos en la hoja comercio_Pesca."""
    try:
        # Flujos y grupos de productos
        for t in range(0, 3):
            hoja.cell(10, 8+t).value = tabla_final.iloc[3, t]
        for t in range(0, 2):
            hoja.cell(10, 13+t).value = tabla_final.iloc[3, t+4]
        
        # Principales destinos
        for x in range(0, 4):
            hoja.cell(37+x, 6).value = _index_label(tabla_destinos.index[x+7])
            for t in range(0, 3):
                hoja.cell(37+x, 8+t).value = tabla_destinos.iloc[x+7, t]
            hoja.cell(37+x, 13).value = tabla_destinos.iloc[x+7, 4]
            hoja.cell(37+x, 14).value = tabla_destinos.iloc[x+7, 5]
        
        # Numero de destinos
        for t in range(0, 3):
            hoja.cell(41, 8+t).value = num_destinos.iloc[1, t]
        
        logger.debug("Hoja comercio_Pesca completada")
    except Exception as e:
        logger.error(f"Error escribiendo comercio_Pesca: {str(e)}")
        raise


def _escribir_indices_textil(hoja, indice_textil):
    """Escribe los indices del sector textil en la hoja Indices_X_Textil."""
    try:
        if not indice_textil:
            logger.warning("No se encontraron indices para Textil")
            return

        bloque_total = indice_textil.get('total', {})
        _limpiar_bloque_indices(hoja, 11, 3010, 1, 5)
        _limpiar_bloque_indices(hoja, 11, 3010, 8, 12)
        _escribir_bloque_indices(hoja, bloque_total.get('TM_sector', pd.DataFrame()), 11, 1)
        _escribir_bloque_indices(hoja, bloque_total.get('FOB_sector', pd.DataFrame()), 11, 8)

        bloques_fob = {
            'confecciones': 27,
            'prendas_vestir': 46,
            'otras_confecciones': 65,
            'textiles': 84,
            'fibras_textiles': 103,
            'tejidos': 122,
            'hilos_hilados': 141,
        }

        for etiqueta, col_fob in bloques_fob.items():
            bloque = indice_textil.get(etiqueta, {})
            _limpiar_bloque_indices(hoja, 11, 3010, col_fob, col_fob + 4)
            _escribir_bloque_indices(hoja, bloque.get('FOB_sector', pd.DataFrame()), 11, col_fob)

        logger.debug("Hoja Indices_X_Textil completada")
    except Exception as e:
        logger.error(f"Error escribiendo Indices_X_Textil: {str(e)}")
        raise


def _escribir_indices_textil_import(hoja, indice_textil):
    """Escribe los indices de importaciones en la hoja Indices_M_Textil."""
    try:
        if not indice_textil:
            logger.warning("No se encontraron indices de importaciones para Textil")
            return

        bloque_total = indice_textil.get('total', {})
        _limpiar_bloque_indices(hoja, 11, 3010, 1, 5)
        _limpiar_bloque_indices(hoja, 11, 3010, 8, 12)
        _escribir_bloque_indices(hoja, bloque_total.get('TM_sector', pd.DataFrame()), 11, 1)
        _escribir_bloque_indices(hoja, bloque_total.get('FOB_sector', pd.DataFrame()), 11, 8)

        bloques_fob = {
            'textiles': 27,
            'tejidos': 46,
            'hilos_hilados': 65,
            'fibras_textiles': 84,
            'confecciones': 103,
            'prendas_vestir': 122,
            'otras_confecciones': 141,
        }

        for etiqueta, col_fob in bloques_fob.items():
            bloque = indice_textil.get(etiqueta, {})
            _limpiar_bloque_indices(hoja, 11, 3010, col_fob, col_fob + 4)
            _escribir_bloque_indices(hoja, bloque.get('FOB_sector', pd.DataFrame()), 11, col_fob)

        logger.debug("Hoja Indices_M_Textil completada")
    except Exception as e:
        logger.error(f"Error escribiendo Indices_M_Textil: {str(e)}")
        raise


def _limpiar_bloque_indices(hoja, fila_inicio, fila_fin, col_inicio, col_fin):
    for fila in range(fila_inicio, fila_fin + 1):
        for col in range(col_inicio, col_fin + 1):
            hoja.cell(fila, col).value = None


def _escribir_bloque_indices(hoja, tabla, fila_inicio, col_inicio):
    if tabla is None or tabla.empty:
        return

    tabla_exportar = tabla.reset_index()

    for fila_offset, (_, row) in enumerate(tabla_exportar.iterrows()):
        fila_excel = fila_inicio + fila_offset
        for col_offset, valor in enumerate(row):
            hoja.cell(fila_excel, col_inicio + col_offset).value = valor

